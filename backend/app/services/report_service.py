import io
import csv
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from backend.app.services.attendance_service import AttendanceService

class ReportService:
    @staticmethod
    def generate_csv(subject: str = None, date: str = None):
        records = AttendanceService.get_attendance_records(date=date, subject=subject)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Enrollment ID", "Student Name", "Subject", "Date", "Time", "Status", "Method"])
        
        for r in records:
            writer.writerow([
                r["enrollment"],
                r["name"],
                r["subject"],
                r["date"],
                r["time"],
                r["status"],
                r["method"]
            ])
            
        output.seek(0)
        return output.getvalue().encode("utf-8")

    @staticmethod
    def generate_excel(subject: str = None, date: str = None):
        records = AttendanceService.get_attendance_records(date=date, subject=subject)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Styling tokens
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        cell_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB")
        )
        
        # Title row
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"AttendAI Attendance Report - {subject or 'All Subjects'} ({date or 'All Dates'})"
        title_cell.font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Table Headers
        headers = ["Enrollment ID", "Student Name", "Subject", "Date", "Time", "Status", "Method"]
        ws.append([]) # Row 2 empty spacer
        ws.append(headers) # Row 3
        
        for col_idx in range(1, 8):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = cell_alignment

        # Data Rows
        for r in records:
            row_data = [
                r["enrollment"],
                r["name"],
                r["subject"],
                r["date"],
                r["time"],
                r["status"],
                r["method"]
            ]
            ws.append(row_data)
            
            row_idx = ws.max_row
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_pdf(subject: str = None, date: str = None):
        records = AttendanceService.get_attendance_records(date=date, subject=subject)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=20,
            textColor=colors.HexColor("#1E3A8A"),
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            "ReportSubTitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            textColor=colors.HexColor("#4B5563"),
            spaceAfter=15
        )

        elements = []
        elements.append(Paragraph("AttendAI Attendance Report", title_style))
        elements.append(Paragraph(f"Subject: <b>{subject or 'All'}</b> &nbsp;|&nbsp; Date: <b>{date or 'All'}</b> &nbsp;|&nbsp; Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        elements.append(Spacer(1, 10))

        # Table data
        table_data = [["Enrollment", "Student Name", "Subject", "Date", "Time", "Status"]]
        for r in records:
            table_data.append([
                r["enrollment"],
                r["name"],
                r["subject"],
                r["date"],
                r["time"],
                r["status"]
            ])

        if len(table_data) == 1:
            table_data.append(["No records found", "-", "-", "-", "-", "-"])

        table = Table(table_data, colWidths=[90, 140, 100, 80, 70, 60])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#1F2937")),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
